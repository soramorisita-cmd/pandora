dir # -*- coding: utf-8 -*-
"""
add_product.py
──────────────
【scanモード】YupooカテゴリURLをスキャンしてブランド別JSONに保存
  python add_product.py scan "https://angelking47.x.yupoo.com/categories/5098470"

【importモード】data/フォルダのJSONを選んでサイトに反映
  python add_product.py import
  python add_product.py import --no-images
  python add_product.py import --no-push
"""

import argparse, json, re, subprocess, time
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote, urlparse, parse_qs, unquote

import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────── 設定 ───────────────────
DATA_DIR      = Path(__file__).parent / "data"
PRODUCTS_JSON = Path(__file__).parent / "products.json"
AFFCODE       = "a235412"
KAKOBUY_BASE  = "https://www.kakobuy.com/item/details"
CNY_TO_JPY    = 21.5
CNY_TO_USD    = 0.138

DATA_DIR.mkdir(exist_ok=True)

PURCHASE_DOMAINS = [
    "taobao.com", "1688.com", "weidian.com",
    "tmall.com", "detail.tmall.com", "pinduoduo.com", "jd.com",
]

# ─────────────────── 商品タイプ分類 ───────────────────
PRODUCT_TYPES = [
    ("SNEAKERS",     ["sneaker","shoe","jordan","dunk","air force","yeezy","kobe","lebron","vomero","samba","campus","ultraboost","foam runner","スニーカー","球鞋","运动鞋","篮球鞋","跑鞋","aj","nike air","new balance"]),
    ("HOODIES",      ["hoodie","hoody","hooded","zip hoodie","连帽卫衣","连帽","パーカ"]),
    ("JACKETS",      ["jacket","coat","shell","puffer","wind","ジャケット","外套","夹克","棉服","羽绒","棒球"]),
    ("SWEATERS",     ["sweater","crewneck","crew neck","knit","sweatshirt","pullover","圆领","卫衣","针织","毛衣"]),
    ("T-SHIRTS",     ["tee","t-shirt","t shirt","tank top","short sleeve","tシャツ","短袖","t恤"]),
    ("TOPS",         ["top","vest","polo","rugby","football","トップス","背心"]),
    ("SHIRTS",       ["shirt","flannel","oxford","button","シャツ","衬衫"]),
    ("SHORTS",       ["short","ショーツ","短裤","sweatshort","soccer short"]),
    ("PANTS",        ["pant","jogger","sweatpant","trouser","denim","jeans","卫裤","长裤","工装裤","运动裤"]),
    ("BAGS",         ["bag","tote","backpack","pack","pouch","バッグ","背包","挎包","包"]),
    ("ACCESSORIES",  ["cap","hat","beanie","sock","glove","keychain","belt","scarf","帽","袜","手套","钥匙","围巾"]),
]

# Yupooサブカテゴリ名 → PANDORAタイプ マッピング（小文字で比較）
SUBCATEGORY_MAP = {
    "tee": "T-SHIRTS", "t-shirt": "T-SHIRTS", "t shirt": "T-SHIRTS",
    "short sleeve": "T-SHIRTS", "jersey": "T-SHIRTS",
    "hoodie": "HOODIES", "hoody": "HOODIES", "zip hoodie": "HOODIES",
    "zip hood": "HOODIES", "hood": "HOODIES",
    "sweater": "SWEATERS", "crewneck": "SWEATERS", "crew": "SWEATERS",
    "sweatshirt": "SWEATERS", "knit": "SWEATERS",
    "jacket": "JACKETS", "coat": "JACKETS", "shell tracksuit": "JACKETS",
    "shell": "JACKETS", "puffer": "JACKETS", "windbreaker": "JACKETS",
    "bottoms": "PANTS", "pants": "PANTS", "jogger": "PANTS",
    "trousers": "PANTS", "jeans": "PANTS",
    "shorts": "SHORTS", "short": "SHORTS", "short tracksuit": "SHORTS",
    "tracksuit": "SWEATERS", "track": "SWEATERS",
    "shirt": "SHIRTS",
    "top": "TOPS", "vest": "TOPS", "polo": "TOPS",
    "bag": "BAGS", "backpack": "BAGS", "tote": "BAGS",
    "accessories": "ACCESSORIES", "wear accessories": "ACCESSORIES",
    "accessory": "ACCESSORIES", "cap": "ACCESSORIES", "hat": "ACCESSORIES",
    "beanie": "ACCESSORIES", "sock": "ACCESSORIES", "glove": "ACCESSORIES",
    "keychain": "ACCESSORIES",
    # スニーカー系
    "sneaker": "SNEAKERS", "sneakers": "SNEAKERS",
    "shoe": "SNEAKERS", "shoes": "SNEAKERS",
    "jordan": "SNEAKERS", "dunk": "SNEAKERS", "force": "SNEAKERS",
    "yeezy": "SNEAKERS", "boost": "SNEAKERS",
    "new balance": "SNEAKERS", "nb": "SNEAKERS",
    "kobe": "SNEAKERS", "lebron": "SNEAKERS",
    "slide": "SNEAKERS", "foam runner": "SNEAKERS",
    "sacai": "SNEAKERS", "vomero": "SNEAKERS",
    "campus": "SNEAKERS", "samba": "SNEAKERS", "sambas": "SNEAKERS",
    "ultraboost": "SNEAKERS", "ultra boost": "SNEAKERS",
    "balenciaga": "SNEAKERS", "runner": "SNEAKERS",
    "air max": "SNEAKERS", "airmax": "SNEAKERS",
}

def map_subcategory(name: str) -> str:
    """サブカテゴリ名をPANDORAタイプに変換（前方一致・部分一致）"""
    nl = name.lower().strip()
    # 完全一致優先
    if nl in SUBCATEGORY_MAP:
        return SUBCATEGORY_MAP[nl]
    # 前方一致（長いキーから試す）
    for key in sorted(SUBCATEGORY_MAP.keys(), key=len, reverse=True):
        if nl.startswith(key) or key in nl:
            return SUBCATEGORY_MAP[key]
    return None  # マッチなし → classify()にフォールバック

TYPE_COLORS = {
    "HOODIES":"FFF2CC","JACKETS":"FFE6CC","SWEATERS":"DDEBF7",
    "T-SHIRTS":"D9EAD3","TOPS":"EAD1DC","SHIRTS":"CCE5FF",
    "SHORTS":"D0E4F5","PANTS":"E6CCFF","BAGS":"D9D9D9",
    "ACCESSORIES":"F4CCCC","Other":"FFFFFF",
}

def classify(title: str) -> str:
    tl = title.lower()
    for ptype, keywords in PRODUCT_TYPES:
        for kw in keywords:
            if kw in tl:
                return ptype
    return "Other"

# ─────────────────── 為替レート ───────────────────
def fetch_rates():
    global CNY_TO_JPY, CNY_TO_USD
    try:
        r = requests.get("https://api.exchangerate-api.com/v4/latest/CNY", timeout=5)
        rates = r.json().get("rates", {})
        CNY_TO_JPY = rates.get("JPY", CNY_TO_JPY)
        CNY_TO_USD = rates.get("USD", CNY_TO_USD)
        print(f"💱 為替: 1CNY = {CNY_TO_JPY:.1f}円 / {CNY_TO_USD:.4f}USD")
    except:
        print(f"⚠ 為替取得失敗 → デフォルト値使用")

# ─────────────────── URL処理 ───────────────────
def extract_purchase_url(url: str) -> str:
    if "yupoo.com/external" in url:
        parsed = urlparse(url)
        raw = parse_qs(parsed.query).get("url", [""])[0]
        return unquote(unquote(raw)).replace("&amp;", "&")
    return url

def to_kakobuy(url: str) -> str:
    return f"{KAKOBUY_BASE}?url={quote(url, safe='')}&affcode={AFFCODE}" if url else ""

def get_base_url(url: str) -> str:
    p = urlparse(url)
    return f"{p.scheme}://{p.netloc}"

def parse_yupoo_url(url: str) -> dict:
    m = re.match(r"https?://([^.]+)\.x\.yupoo\.com/(categories|albums)/(\d+)", url)
    if not m:
        raise ValueError(f"YupooのURLではありません: {url}")
    return {"seller": m.group(1), "url_type": m.group(2), "id": m.group(3)}

def brand_to_filename(brand: str) -> str:
    safe = re.sub(r'[\\/*?:"<>|]', "", brand).strip().replace(" ", "_")
    return safe + ".json"

# ─────────────────── スクレイパー ───────────────────
def extract_batch_from_title(title: str) -> str | None:
    """
    タイトルから【バッチ名】を抽出する。
    バッチ名: PK / LJR / OG / DT / M版 / TOP など（短い英字 or 日本語）
    除外: 価格（350yuan等）/ 商品番号（FD2629-100等）/ Pre-order等
    """
    brackets = re.findall(r'【([^】]+)】', title)
    for b in brackets:
        b = b.strip()
        # 価格・数字のみはスキップ
        if re.match(r'^[\d,yuan元Y¥￥\s]+$', b, re.IGNORECASE):
            continue
        # Pre-order などスキップ
        if re.search(r'pre|order|sale', b, re.IGNORECASE):
            continue
        # 商品番号パターンをスキップ（例: FD2629-100, BV1310-337, CT0856-600）
        # 英字2〜3文字 + 数字4〜6文字 + ハイフン + 数字
        if re.match(r'^[A-Z]{1,3}\d{4,6}(-\d{1,4})?$', b):
            continue
        # 長すぎるもの（15文字超）はスキップ
        if len(b) > 15:
            continue
        return b
    return None

def scrape_subcategories(category_url: str, ctx) -> list[dict]:
    """
    カテゴリページのメインコンテンツ内にあるバッチボタンを取得する。
    referrercate=親ID で絞り込み、現在のカテゴリの子のみ取得。
    """
    base_url = get_base_url(category_url)
    parent_id_m = re.search(r"/categories/(\d+)", category_url)
    if not parent_id_m:
        return []
    parent_id = parent_id_m.group(1)

    page = ctx.new_page()
    try:
        page.goto(category_url, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)

        # referrercate=親ID を含むリンクのみ取得（現在カテゴリの子だけ）
        js = """
        (parentId) => {
            const results = [];
            const seen = new Set();
            const needle = 'referrercate=' + parentId;
            const links = Array.from(document.querySelectorAll('a[href]'));
            for (const a of links) {
                const href = a.getAttribute('href') || '';
                if (!href.includes(needle)) continue;
                if (!href.includes('isSubCate=true')) continue;
                const name = a.textContent.trim();
                if (!name || seen.has(href)) continue;
                seen.add(href);
                results.push({ name, href });
            }
            return results;
        }
        """
        results = page.evaluate(js, parent_id)
        print(f"   バッチボタン検出: {len(results)}件")
        for r in results[:5]:
            print(f"   → {r['name']}")
    except Exception as e:
        print(f"  ⚠ バッチ取得エラー: {e}")
        results = []
    finally:
        page.close()

    subcats = []
    seen = set()
    for item in results:
        href = item.get("href", "")
        name = item.get("name", "").strip()
        if not name or href in seen:
            continue
        seen.add(href)
        full_url = (base_url + href) if href.startswith("/") else href
        ptype = map_subcategory(name)
        subcats.append({"name": name, "url": full_url, "type": ptype})

    return subcats

def scrape_albums(category_url: str, ctx) -> list[dict]:
    albums = []
    base_url = get_base_url(category_url)
    page_num = 1
    while True:
        url  = f"{category_url}?page={page_num}" if page_num > 1 else category_url
        page = ctx.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)
            html = page.content()
        finally:
            page.close()
        soup  = BeautifulSoup(html, "html.parser")
        links = soup.find_all("a", href=re.compile(r"/albums/\d+"))
        if not links:
            break
        seen = set()
        items = []
        for a in links:
            href = a.get("href", "")
            m    = re.search(r"/albums/(\d+)", href)
            if not m or m.group(1) in seen:
                continue
            seen.add(m.group(1))
            title   = a.get("title", "") or a.get_text(strip=True) or ""
            alb_url = base_url + href if href.startswith("/") else href
            items.append({"album_id": m.group(1), "title": title, "yupoo_url": alb_url})
        if not items:
            break
        albums.extend(items)
        print(f"   ページ{page_num}: {len(items)}件 (累計 {len(albums)}件)")
        pages_text = soup.find(string=re.compile(r"共\d+页"))
        if pages_text:
            total = re.search(r"共(\d+)页", str(pages_text))
            if total and page_num >= int(total.group(1)):
                break
        else:
            break
        page_num += 1
        if page_num > 20:
            break
    return albums

def scrape_taobao_link(album_url: str, ctx, retries=3):
    for attempt in range(1, retries+1):
        page = ctx.new_page()
        try:
            page.goto(album_url, wait_until="domcontentloaded", timeout=40000)
            time.sleep(2)
            html = page.content()
            soup = BeautifulSoup(html, "html.parser")
            for a in soup.find_all("a", href=True):
                h = a["href"]
                if "yupoo.com/external" in h:
                    return extract_purchase_url(h)
                if any(d in h for d in PURCHASE_DOMAINS):
                    return extract_purchase_url(h)
            pat = re.compile(
                r'https?://[^\s\'"<>]*(?:' +
                '|'.join(d.replace('.', r'\.') for d in PURCHASE_DOMAINS) +
                r')[^\s\'"<>]*'
            )
            found = pat.findall(soup.get_text())
            if found:
                return extract_purchase_url(found[0])
            hrefs = page.eval_on_selector_all("a[href]", "els => els.map(e => e.href)")
            for h in hrefs:
                if "yupoo.com/external" in h:
                    return extract_purchase_url(h)
                if any(d in h for d in PURCHASE_DOMAINS):
                    return extract_purchase_url(h)
            return None
        except:
            if attempt < retries:
                time.sleep(attempt * 3)
        finally:
            page.close()
    return None


def scrape_purchase_product_name(purchase_url: str, ctx) -> str:
    """Taobao/Weidian/1688のページから商品名を取得してOther再分類に使用"""
    if not purchase_url:
        return ""
    page = ctx.new_page()
    try:
        page.goto(purchase_url, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")

        # ① <title>タグから取得（最も信頼性が高い）
        title_el = soup.find("title")
        if title_el:
            title = title_el.get_text(strip=True)
            # サイト名部分を除去（例: "商品名-淘宝网" → "商品名"）
            for sep in [" - 淘宝网", "-淘宝网", " - 微店", "- 微店", " | 1688"]:
                title = title.split(sep)[0]
            if len(title) > 3:
                return title.strip()

        # ② og:title メタタグから取得
        og = soup.find("meta", property="og:title")
        if og and og.get("content"):
            return og["content"].strip()

        return ""
    except:
        return ""
    finally:
        page.close()


def reclassify_other(album: dict, ctx) -> str:
    """
    Otherに分類された商品をTaobao/Weidianの商品名で再分類する。
    既存カテゴリにマッチしない場合はOtherのまま返す。
    """
    purchase_url = album.get("purchase", "")
    if not purchase_url:
        return "Other"

    product_name = scrape_purchase_product_name(purchase_url, ctx)
    if not product_name:
        return "Other"

    new_type = classify(product_name)
    if new_type != "Other":
        return new_type

    return "Other"

def scrape_first_image(album_url: str, ctx) -> str:
    """アルバムの1枚目画像URLを取得"""
    page = ctx.new_page()
    try:
        page.goto(album_url, wait_until="networkidle", timeout=40000)
        time.sleep(2)

        # JS実行後のsrcを取得（lazy load対応）
        hrefs = page.eval_on_selector_all(
            "img",
            "els => els.map(e => e.getAttribute('data-src') || e.src || '')"
        )
        for src in hrefs:
            if src and ("photo.yupoo" in src or "uvd.yupoo" in src):
                return src

        # フォールバック: BeautifulSoup
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")
        for img in soup.find_all("img"):
            src = img.get("data-src") or img.get("src") or ""
            if "photo.yupoo" in src or "uvd.yupoo" in src:
                return src
        return ""
    except:
        return ""
    finally:
        page.close()

# ─────────────────── ブランドJSON ───────────────────
def load_brand_json(brand: str) -> dict:
    path = DATA_DIR / brand_to_filename(brand)
    if path.exists():
        return json.loads(path.read_text("utf-8"))
    return {"brand": brand, "seller": "", "updated": "", "products": []}

def save_brand_json(data: dict):
    path = DATA_DIR / brand_to_filename(data["brand"])
    data["updated"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")
    print(f"💾 JSON保存: {path.name}")

# ─────────────────── Excel出力 ───────────────────
THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HEADERS    = ["No.", "商品タイプ", "アルバムタイトル", "YupooURL", "TaobaoURL", "KakobuyURL", "価格(CNY)", "価格(JPY)"]
COL_WIDTHS = [5, 14, 50, 40, 45, 65, 10, 12]

def save_brand_excel(data: dict, path: Path):
    products = data["products"]
    wb = Workbook()
    ws = wb.active
    ws.title = "全件"
    ws.append(HEADERS)
    ws.row_dimensions[1].height = 22
    for c, (_, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(1, c)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, p in enumerate(products, 1):
        color = TYPE_COLORS.get(p.get("type", "Other"), "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        vals  = [i, p.get("type","Other"), p["title"], p["yupoo_url"],
                 p.get("purchase") or "未取得",
                 p.get("kakobuy") or "未取得",
                 p.get("price_cny") or "―",
                 p.get("price_jpy") or "―"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i+1, c, v)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.fill = fill
            cell.border = BORDER
    ws.freeze_panes = "A2"

    groups = defaultdict(list)
    for p in products:
        groups[p.get("type", "Other")].append(p)

    for ptype, items in sorted(groups.items()):
        ws2 = wb.create_sheet(ptype[:31])
        ws2.append(HEADERS)
        ws2.row_dimensions[1].height = 22
        for c, (_, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws2.cell(1, c)
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            ws2.column_dimensions[get_column_letter(c)].width = w
        color = TYPE_COLORS.get(ptype, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for i, p in enumerate(items, 1):
            vals = [i, p.get("type","Other"), p["title"], p["yupoo_url"],
                    p.get("purchase") or "未取得",
                    p.get("kakobuy") or "未取得",
                    p.get("price_cny") or "―",
                    p.get("price_jpy") or "―"]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(i+1, c, v)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center")
                cell.fill = fill
                cell.border = BORDER
        ws2.freeze_panes = "A2"

    wb.save(str(path))
    print(f"📊 Excel保存: {path.name}")

# ─────────────────── products.json（サイト用）───────────────────
def load_site_products() -> list[dict]:
    if PRODUCTS_JSON.exists():
        return json.loads(PRODUCTS_JSON.read_text("utf-8")).get("products", [])
    return []

def save_site_products(products: list[dict]):
    out = {"updated": time.strftime("%Y-%m-%dT%H:%M:%S"), "count": len(products), "products": products}
    PRODUCTS_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), "utf-8")

def merge_to_site(existing: list, new_items: list) -> tuple:
    urls  = {p["yupoo"] for p in existing}
    added = 0
    for item in new_items:
        if item["yupoo"] not in urls:
            existing.append(item)
            urls.add(item["yupoo"])
            added += 1
    return existing, added

def git_push(msg: str):
    repo = Path(__file__).parent
    for cmd in [["git","add","products.json"],["git","commit","-m",msg],["git","push"]]:
        r = subprocess.run(cmd, cwd=repo, capture_output=True, text=True)
        ok = r.returncode == 0 or "nothing to commit" in r.stdout
        print(f"  {'✓' if ok else '⚠'} {' '.join(cmd)}")

# ═══════════════════════════════════════════
#  SCAN モード
# ═══════════════════════════════════════════
def cmd_scan(url: str):
    info   = parse_yupoo_url(url)
    seller = info["seller"]

    brand = input("ブランド名を入力してください > ").strip()
    if not brand:
        print("ブランド名が空です")
        return

    model = input("モデル名を入力してください（スニーカーなど: AIR JORDAN 1 / スキップはEnter） > ").strip() or None
    if model:
        print(f"   モデル: {model} → 同モデルの商品はカタログで1枚のカードにグループ表示されます")

    fetch_rates()
    print(f"\n🔍 スキャン: {seller} / ブランド: {brand}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )

        print("📂 サブカテゴリ確認中...")
        subcats = scrape_subcategories(url, ctx)

        if subcats:
            print(f"✅ バッチ検出: {len(subcats)}件")
            for sc in subcats:
                print(f"   {sc['name']:<25} → {sc['type'] or 'classify()'}")
            print()
            albums = []
            for sc in subcats:
                print(f"📂 [{sc['name']}] スキャン中...")
                sc_albums = scrape_albums(sc["url"], ctx)
                for a in sc_albums:
                    a["type"]  = sc["type"] or classify(a["title"])
                    a["batch"] = sc["name"]  # バッチ名はサブカテゴリ名から確実に取得
                albums.extend(sc_albums)
                print(f"   → {len(sc_albums)}件")
        else:
            print("📂 アルバム一覧取得中...")
            albums = scrape_albums(url, ctx)
            if not albums:
                print("❌ アルバムが取得できませんでした")
                browser.close()
                return
            for a in albums:
                a["type"] = classify(a["title"])

        if not albums:
            print("❌ アルバムが取得できませんでした")
            browser.close()
            return

        # 共通フィールド初期化 + タイトルからバッチを自動抽出
        for a in albums:
            a.setdefault("purchase",  None)
            a.setdefault("kakobuy",   None)
            a.setdefault("price_cny", None)
            a.setdefault("price_jpy", None)
            a.setdefault("image",     None)
            if model:
                a["model"] = model
            # バッチをタイトルから自動抽出（未設定の場合のみ）
            if not a.get("batch"):
                a["batch"] = extract_batch_from_title(a.get("title", ""))

        groups = defaultdict(list)
        for a in albums:
            groups[a["type"]].append(a)
        print(f"\n📊 自動分類結果（{len(albums)}件）:")
        for ptype, items in sorted(groups.items()):
            print(f"   {ptype:<14}: {len(items)}件")

        print(f"\n🔗 購入リンク取得中...")
        for i, alb in enumerate(albums):
            print(f"   [{i+1:03d}/{len(albums)}] {alb['title'][:45]}", end="\r")
            taobao = scrape_taobao_link(alb["yupoo_url"], ctx)
            alb["purchase"] = taobao or ""
            alb["kakobuy"]  = to_kakobuy(taobao) if taobao else ""
            # 画像も同時取得（アルバムページは既に開いているので追加コスト小）
            if not alb.get("image"):
                alb["image"] = scrape_first_image(alb["yupoo_url"], ctx)
            time.sleep(0.8)
        print()

        # Other再分類（Taobao/Weidianの商品名で再試行）
        others = [a for a in albums if a["type"] == "Other" and a.get("purchase")]
        if others:
            print(f"\n🔄 Other再分類中... ({len(others)}件)")
            reclassified = 0
            for i, alb in enumerate(others):
                print(f"   [{i+1:03d}/{len(others)}] {alb['title'][:40]}", end="\r")
                new_type = reclassify_other(alb, ctx)
                if new_type != "Other":
                    print(f"   ✅ [{new_type}] {alb['title'][:40]}")
                    alb["type"] = new_type
                    reclassified += 1
                time.sleep(0.8)
            print(f"\n   再分類成功: {reclassified}件 / Other残り: {len(others)-reclassified}件")

        ctx.close()
        browser.close()

    # ブランドJSONに保存
    data = load_brand_json(brand)
    data["brand"]  = brand
    data["seller"] = seller

    existing_ids = {p.get("album_id") for p in data["products"]}
    added = sum(1 for a in albums if a["album_id"] not in existing_ids)
    for a in albums:
        if a["album_id"] not in existing_ids:
            data["products"].append(a)

    save_brand_json(data)

    excel_path = DATA_DIR / brand_to_filename(brand).replace(".json", ".xlsx")
    save_brand_excel(data, excel_path)

    print(f"\n✅ 完了！")
    print(f"   新規: {added}件 / 合計: {len(data['products'])}件")
    print(f"   ファイル: data/{brand_to_filename(brand)}")
    print(f"\n📝 次のステップ:")
    print(f"   1. data/{excel_path.name} で分類を確認・修正")
    print(f"   2. python add_product.py import でサイトに反映")

# ═══════════════════════════════════════════
#  IMPORT モード
# ═══════════════════════════════════════════
def cmd_import(no_images: bool, no_push: bool):
    json_files = sorted(DATA_DIR.glob("*.json"))
    if not json_files:
        print(f"❌ data/ フォルダにJSONファイルがありません")
        return

    print(f"\n📂 利用可能なブランドデータ:")
    for i, f in enumerate(json_files, 1):
        data  = json.loads(f.read_text("utf-8"))
        count = len(data.get("products", []))
        print(f"  [{i}] {f.stem:<25} {count}件")
    print(f"  [A] 全て反映")

    sel = input("\n番号を選択（複数はカンマ区切り: 1,3）> ").strip().upper()

    if sel == "A":
        selected = json_files
    else:
        indices  = [int(x.strip())-1 for x in sel.split(",") if x.strip().isdigit()]
        selected = [json_files[i] for i in indices if 0 <= i < len(json_files)]

    if not selected:
        print("❌ 無効な選択")
        return

    print(f"\n📥 {len(selected)}ブランドをインポートします")
    all_new = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )

        for f in selected:
            data     = json.loads(f.read_text("utf-8"))
            brand    = data["brand"]
            products = data["products"]
            print(f"\n🏷  {brand} ({len(products)}件)")

            for i, p in enumerate(products):
                print(f"  [{i+1:03d}/{len(products)}] {p['title'][:50]}", end="\r")
                if not no_images and not p.get("image"):
                    p["image"] = scrape_first_image(p["yupoo_url"], ctx)
                    time.sleep(0.5)

                all_new.append({
                    "seller":    data.get("seller", ""),
                    "brand":     brand,
                    "type":      p.get("type", "Other"),
                    "title":     p["title"],
                    "yupoo":     p["yupoo_url"],
                    "purchase":  p.get("purchase", ""),
                    "kakobuy":   p.get("kakobuy", ""),
                    "image":     p.get("image", ""),
                    "price_cny": p.get("price_cny"),
                    "price_jpy": p.get("price_jpy"),
                })

            data["products"] = products
            save_brand_json(data)
            print()

        ctx.close()
        browser.close()

    existing      = load_site_products()
    merged, added = merge_to_site(existing, all_new)
    save_site_products(merged)
    print(f"\n✅ {added}件追加（合計 {len(merged)}件）→ products.json")

    if added == 0:
        print("（全て重複のためスキップ）")
        return

    if not no_push:
        brands = ", ".join(f.stem for f in selected)
        print("\n📤 Git push...")
        git_push(f"add {brands}")
        print(f"\n🌐 約30秒後に反映: https://pandora-doj.pages.dev")
    else:
        print("\n反映するには: git add products.json && git commit -m 'update' && git push")

# ═══════════════════════════════════════════
#  エントリーポイント
# ═══════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Pandora catalog tool")
    sub    = parser.add_subparsers(dest="mode")

    p_scan = sub.add_parser("scan", help="YupooURLをスキャンしてブランド別JSONに保存")
    p_scan.add_argument("url", help="YupooカテゴリURL")

    p_imp = sub.add_parser("import", help="data/のJSONを選んでサイトに反映")
    p_imp.add_argument("--no-images", action="store_true")
    p_imp.add_argument("--no-push",   action="store_true")

    args = parser.parse_args()

    if args.mode == "scan":
        cmd_scan(args.url)
    elif args.mode == "import":
        cmd_import(getattr(args,"no_images",False), getattr(args,"no_push",False))
    else:
        parser.print_help()

if __name__ == "__main__":
    main()